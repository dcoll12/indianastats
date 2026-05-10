"""
Build Google Sheets county-candidates workbook — single-sheet, click-to-select edition.

All controls on one sheet ("Lookup"):
  Left panel  (cols B-E):  county names — CLICK a name to select/deselect it.
                            Selected counties highlight gold.
  Filter bar  (row 4):     District Type dropdown + District # text input.
  Right panel (cols G-M):  results via QUERY formula with dynamic WHERE clause.

Formula fix: replaced SORT(FILTER(…), {1,3}, {TRUE,TRUE}) — which is Excel
syntax that Google Sheets rejects — with a QUERY string built dynamically
via LET + TEXTJOIN so all three filters compose cleanly.

Click-to-select: onSelectionChange in Apps Script toggles the hidden state
cell (col C or E) whenever the user clicks a county name cell (col B or D).
Conditional formatting lights up the county name gold when selected.

Other sheets:
  By District  – flat sorted reference table (auto-filter)
  All Data     – hidden source (winners / uncontested only, 736 rows)
  Apps Script  – full script + setup instructions
"""

import csv, json, os, re, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
BLUE_DARK   = "0C2340"
BLUE_MID    = "1E4488"
GOLD        = "C8962E"
GOLD_LIGHT  = "F5DFA0"
WHITE       = "FFFFFF"
GREY_LIGHT  = "F2F4F7"
GREY_MED    = "D0D5DD"
RED_LIGHT   = "FFE4E4"
BLUE_LIGHT  = "E4EEFF"
GREEN_LIGHT = "E4F5E4"
WIN_FILL    = "D4EDDA"

def thin(top=True, bottom=True, left=True, right=True):
    s = {n: Side(style="thin", color=GREY_MED) if f else Side(style=None)
         for n, f in [("top",top),("bottom",bottom),("left",left),("right",right)]}
    return Border(**s)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
base          = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESULTS_CSV   = os.path.join(base, "AllOfficeResults(3).csv")
CANDS_JSON    = os.path.join(base, "data", "candidates_2026.json")
DISTRICT_JSON = os.path.join(base, "data", "district_data.json")

with open(CANDS_JSON)    as f: cands_json = json.load(f)
with open(DISTRICT_JSON) as f: dm         = json.load(f)

# ---------------------------------------------------------------------------
# Parse CSV → determine primary winners
# ---------------------------------------------------------------------------
ORDINALS = {"first":1,"second":2,"third":3,"fourth":4,"fifth":5,
            "sixth":6,"seventh":7,"eighth":8,"ninth":9}

def office_to_key(office):
    o = office.strip()
    if "State Representative" in o:
        m = re.search(r"District\s+(\d+)", o)
        return ("hd", int(m.group(1))) if m else None
    if "State Senator" in o:
        m = re.search(r"District\s+(\d+)", o)
        return ("sd", int(m.group(1))) if m else None
    if "United States Representative" in o:
        m = re.search(r"(\w+)\s+District", o, re.I)
        if m:
            w = m.group(1).lower()
            return ("cd", int(w) if w.isdigit() else ORDINALS.get(w))
    return None

vote_totals = collections.defaultdict(int)
seats_map   = {}
with open(RESULTS_CSV, newline="", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        key = office_to_key(row["Office"])
        if not key: continue
        dtype, dnum = key
        party = row["PoliticalParty"].strip()
        name  = row["NameonBallot"].strip()
        vote_totals[(dtype, dnum, party, name)] += int(row["TotalVotes"] or 0)
        seats_map[(dtype, dnum, party)] = int(row["NumberofOfficeSeats"] or 1)

norm     = lambda s: re.sub(r"\s+", " ", s.lower().strip())
winners  = set()
vote_map = {}
by_race  = collections.defaultdict(list)
for (dtype, dnum, party, name), votes in vote_totals.items():
    by_race[(dtype, dnum, party)].append((name, votes))
for (dtype, dnum, party), candidates in by_race.items():
    n = seats_map.get((dtype, dnum, party), 1)
    for cname, _ in sorted(candidates, key=lambda x: -x[1])[:n]:
        winners.add((dtype, dnum, norm(cname)))
    for cname, votes in candidates:
        vote_map[(dtype, dnum, norm(cname))] = votes

def candidate_result(dtype, dnum, name):
    race_has_results = any(
        (dtype, dnum, p) in seats_map
        for p in ("Democratic", "Republican", "Libertarian", "Independent")
    )
    if not race_has_results:
        return ("Uncontested", 2, 0)
    nkey  = (dtype, dnum, norm(name))
    votes = vote_map.get(nkey, 0)
    if votes == 0:
        return ("Uncontested", 2, 0)
    return ("Won Primary ✓", 0, votes) if nkey in winners else None

# ---------------------------------------------------------------------------
# Build flat rows (winners + uncontested only)
# ---------------------------------------------------------------------------
rows = []
for county in sorted(dm["all_counties"]):
    for hd in sorted(dm["county_to_hds"].get(county, [])):
        for c in cands_json.get("hd", {}).get(str(hd), []):
            r = candidate_result("hd", hd, c["name"])
            if r: rows.append((county, 1, "House", hd, f"HD-{hd}", c["name"], c["party"], *r))
    for sd in sorted(dm["county_to_sds"].get(county, [])):
        for c in cands_json.get("sd", {}).get(str(sd), []):
            r = candidate_result("sd", sd, c["name"])
            if r: rows.append((county, 2, "Senate", sd, f"SD-{sd}", c["name"], c["party"], *r))
    for cd in sorted(dm["county_to_cds"].get(county, [])):
        for c in cands_json.get("cd", {}).get(str(cd), []):
            r = candidate_result("cd", cd, c["name"])
            if r: rows.append((county, 3, "Congressional", cd, f"CD-{cd}", c["name"], c["party"], *r))

all_counties  = sorted(dm["all_counties"])
half          = 46        # 46 left + 46 right = 92
CB_START      = 6         # first county row
CB_END        = CB_START + half - 1   # 51
last_data_row = len(rows) + 1
print(f"Rows: {len(rows)}")

# ============================================================
wb = Workbook()
wb.remove(wb.active)

# ============================================================
# Sheet: All Data  (hidden)
# A=County  B=SortOrder  C=Type  D=Num  E=District
# F=Candidate  G=Party  H=Result  I=ResultSort  J=Votes
# ============================================================
ws_data = wb.create_sheet("All Data")
ws_data.sheet_state = "hidden"
DATA_HDRS = ["County","SortOrder","Type","Num","District",
             "Candidate","Party","Result","ResultSort","Votes"]
ws_data.append(DATA_HDRS)
for r in rows:
    ws_data.append(list(r))
for col in range(1, len(DATA_HDRS)+1):
    c = ws_data.cell(1, col)
    c.font      = Font(bold=True, color=WHITE, size=10)
    c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center")
for i, w in enumerate([18,10,16,6,10,35,8,18,10,12], 1):
    ws_data.column_dimensions[ws_data.cell(1,i).column_letter].width = w

# ============================================================
# Sheet: Lookup  (single interactive sheet)
# ============================================================
ws = wb.create_sheet("Lookup", 0)

# ---- Row 1: Banner ----
ws.merge_cells("A1:M1")
ws["A1"].value     = "Indiana 2026 Primary Winners — Candidate Lookup"
ws["A1"].font      = Font(bold=True, size=16, color=WHITE)
ws["A1"].fill      = PatternFill("solid", fgColor=BLUE_DARK)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 38

# ---- Row 2: Subtitle ----
ws.merge_cells("A2:M2")
ws["A2"].value     = ("Click a county name to select or deselect it — selected counties highlight gold.  "
                      "Leave all unselected to show every county.  "
                      "Use the Type and Dist # fields to filter by district.")
ws["A2"].font      = Font(size=10, italic=True, color=BLUE_DARK)
ws["A2"].fill      = PatternFill("solid", fgColor=GOLD_LIGHT)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 28

ws.row_dimensions[3].height = 6

# ---- Row 4: Filter controls ----
ws.row_dimensions[4].height = 28

ws["B4"].value     = "Counties:"
ws["B4"].font      = Font(bold=True, size=11, color=BLUE_DARK)
ws["B4"].alignment = Alignment(horizontal="right", vertical="center")

# County count (spans C4:E4 for visual width, actual formula in C4)
ws.merge_cells("C4:E4")
ws["C4"].value = (
    # COUNTIF(...,TRUE) + COUNTIF(...,"TRUE") catches both boolean and text forms
    "=LET(n,COUNTIF(C6:C51,TRUE)+COUNTIF(C6:C51,\"TRUE\")"
    "+COUNTIF(E6:E51,TRUE)+COUNTIF(E6:E51,\"TRUE\"),"
    'IF(n=0,"All 92 counties — click to filter",n&" of 92 selected"))'
)
ws["C4"].font      = Font(bold=True, size=11, color=BLUE_DARK)
ws["C4"].fill      = PatternFill("solid", fgColor=GOLD_LIGHT)
ws["C4"].alignment = Alignment(horizontal="center", vertical="center")
ws["C4"].border    = thin()

ws["H4"].value     = "District Type:"
ws["H4"].font      = Font(bold=True, size=11, color=BLUE_DARK)
ws["H4"].alignment = Alignment(horizontal="right", vertical="center")

ws["I4"].value     = "All Types"
ws["I4"].font      = Font(bold=True, size=11, color=BLUE_DARK)
ws["I4"].fill      = PatternFill("solid", fgColor=GOLD_LIGHT)
ws["I4"].alignment = Alignment(horizontal="center", vertical="center")
ws["I4"].border    = thin()

dv_type = DataValidation(
    type="list",
    formula1='"All Types,House,Senate,Congressional"',
    allow_blank=False, showDropDown=False,
)
dv_type.sqref = "I4"
ws.add_data_validation(dv_type)

ws["K4"].value     = "District #:"
ws["K4"].font      = Font(bold=True, size=11, color=BLUE_DARK)
ws["K4"].alignment = Alignment(horizontal="right", vertical="center")

ws["L4"].value     = ""
ws["L4"].font      = Font(bold=True, size=11, color=BLUE_DARK)
ws["L4"].fill      = PatternFill("solid", fgColor=GOLD_LIGHT)
ws["L4"].alignment = Alignment(horizontal="center", vertical="center")
ws["L4"].border    = thin()

# Showing count (G4)
ws["G4"].value     = '=IFERROR(IF(COUNTA(G6:G5000)=0,"","Showing "&COUNTA(G6:G5000)&" candidates"),"")'
ws["G4"].font      = Font(bold=True, size=10, color=BLUE_DARK)
ws["G4"].alignment = Alignment(horizontal="center", vertical="center")

# ---- Row 5: Column headers ----
ws.row_dimensions[5].height = 22

for col_letter, hdr, bg in [
    ("B","County",BLUE_MID), ("C","",BLUE_MID),
    ("D","County",BLUE_MID), ("E","",BLUE_MID),
    ("G","District",BLUE_MID), ("H","Type",BLUE_MID), ("I","Dist #",BLUE_MID),
    ("J","Candidate Name",BLUE_MID), ("K","Party",BLUE_MID),
    ("L","Primary Result",BLUE_MID), ("M","Votes",BLUE_MID),
]:
    c = ws[f"{col_letter}5"]
    c.value     = hdr
    c.font      = Font(bold=True, size=10, color=WHITE)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin()

# ---- Rows 6-51: County names + hidden state cells ----
# B = county name left   C = state (TRUE/FALSE, narrow)
# D = county name right  E = state (TRUE/FALSE, narrow)
dv_check = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
dv_check.sqref = f"C{CB_START}:C{CB_END} E{CB_START}:E{CB_END}"
ws.add_data_validation(dv_check)

for i, county in enumerate(all_counties):
    row = CB_START + (i % half)
    bg  = GREY_LIGHT if (i % half) % 2 == 0 else WHITE

    if i < half:
        name_col, state_col = 2, 3   # B, C
    else:
        name_col, state_col = 4, 5   # D, E

    nc = ws.cell(row, name_col)
    nc.value     = county
    nc.font      = Font(size=10, color=BLUE_DARK)
    nc.fill      = PatternFill("solid", fgColor=bg)
    nc.alignment = Alignment(vertical="center", indent=1)
    nc.border    = thin()

    sc = ws.cell(row, state_col)
    sc.value     = False
    sc.fill      = PatternFill("solid", fgColor=bg)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    sc.border    = thin()

for row in range(CB_START, CB_END + 1):
    ws.row_dimensions[row].height = 18

# ---- Conditional formatting: selected counties light up gold ----
# Use OR of boolean TRUE and text "TRUE" — Google Sheets imports booleans
# from xlsx but Apps Script setValue() always writes booleans going forward.
ws.conditional_formatting.add(
    f"B{CB_START}:B{CB_END}",
    FormulaRule(
        formula=["($C6=TRUE)+($C6=\"TRUE\")>0"],
        fill=PatternFill("solid", fgColor=GOLD_LIGHT),
        font=Font(color=BLUE_DARK, bold=True, size=10),
    ),
)
ws.conditional_formatting.add(
    f"D{CB_START}:D{CB_END}",
    FormulaRule(
        formula=["($E6=TRUE)+($E6=\"TRUE\")>0"],
        fill=PatternFill("solid", fgColor=GOLD_LIGHT),
        font=Font(color=BLUE_DARK, bold=True, size=10),
    ),
)

# ---- FILTER formula in G6 ----
#
# County check uses MMULT (matrix multiply) — the most reliable approach
# in Google Sheets for "is this row's county in the selected list?":
#
#   match_matrix  =  736×46  boolean: data_county_i == left_county_j
#   selection_vec =  46×1    numeric: 1 if county j is selected, else 0
#   MMULT(match_matrix, selection_vec)  →  736×1: # selected counties matching
#   result > 0  →  county is selected
#
# When nothing is selected, use All Data col B (SortOrder, always 1/2/3 > 0)
# as an all-TRUE stand-in so every row passes.
#
# All Data is already ordered House→Senate→CD within each county, so no SORT
# is needed — rows come out in the correct order from FILTER.
#
# Handles xlsx-imported text "TRUE"/"FALSE" AND boolean true/false from
# Apps Script by summing both comparisons in ls/rs.
L  = last_data_row        # 737  (header row 1 + 736 data rows → data in 2:{L})
CB_S = CB_START           # 6
CB_E = CB_END             # 51

filter_formula = (
    "=IFERROR("
    "LET("
    # ls/rs: 46×1 numeric — 1 if selected (text OR boolean TRUE), else 0
    f"ls,(C{CB_S}:C{CB_E}=TRUE)+(C{CB_S}:C{CB_E}=\"TRUE\"),"
    f"rs,(E{CB_S}:E{CB_E}=TRUE)+(E{CB_S}:E{CB_E}=\"TRUE\"),"
    "n,SUMPRODUCT(ls)+SUMPRODUCT(rs),"
    # sl/sr: replace unselected county names with sentinel "~~~" so they never match
    f"sl,IF(ls>0,B{CB_S}:B{CB_E},\"~~~\"),"
    f"sr,IF(rs>0,D{CB_S}:D{CB_E},\"~~~\"),"
    # co: 736×1 county match — BYROW checks each data row's county against selected lists
    # BYROW+SUMPRODUCT is more reliable in Google Sheets than MMULT 2-D broadcasting
    f"co,IF(n=0,SEQUENCE({L-1})>0,"
    f"BYROW('All Data'!A2:A{L},LAMBDA(x,"
    "SUMPRODUCT(--(x=sl))+SUMPRODUCT(--(x=sr))>0))),"
    # to: 736×1 type filter
    f"to,(I4=\"All Types\")+('All Data'!C2:C{L}=I4)>0,"
    # do: 736×1 district-number filter
    f"do,(LEN(TRIM(L4))=0)+('All Data'!D2:D{L}=IFERROR(VALUE(L4),0))>0,"
    # return matching rows; cols: District,Type,Num,Candidate,Party,Result,Votes
    f"FILTER({{'All Data'!E2:E{L},'All Data'!C2:C{L},'All Data'!D2:D{L},"
    f"'All Data'!F2:F{L},'All Data'!G2:G{L},'All Data'!H2:H{L},'All Data'!J2:J{L}}},"
    "co*to*do)"
    "),"
    "\"No candidates match the selected filters.\""
    ")"
)
ws["G6"].value = filter_formula

# ---- Conditional formatting: right panel ----
ws.conditional_formatting.add("K6:K5000",
    FormulaRule(formula=['$K6="D"'],
                fill=PatternFill("solid", fgColor=RED_LIGHT),
                font=Font(color="8B0000", bold=True)))
ws.conditional_formatting.add("K6:K5000",
    FormulaRule(formula=['$K6="R"'],
                fill=PatternFill("solid", fgColor=BLUE_LIGHT),
                font=Font(color="00008B", bold=True)))
ws.conditional_formatting.add("K6:K5000",
    FormulaRule(formula=['AND($K6<>"D",$K6<>"R",$K6<>"")'],
                fill=PatternFill("solid", fgColor=GREEN_LIGHT),
                font=Font(color="005500", bold=True)))
ws.conditional_formatting.add("L6:L5000",
    FormulaRule(formula=['LEFT($L6,3)="Won"'],
                fill=PatternFill("solid", fgColor=WIN_FILL),
                font=Font(color="155724", bold=True)))

# ---- Column widths ----
col_widths = {
    "A": 1,   # margin
    "B": 19,  # county name left
    "C": 3,   # state col left (narrow — just stores TRUE/FALSE)
    "D": 19,  # county name right
    "E": 3,   # state col right
    "F": 2,   # divider
    "G": 11,  # District label
    "H": 14,  # Type
    "I": 8,   # Dist #
    "J": 36,  # Candidate Name
    "K": 7,   # Party
    "L": 18,  # Primary Result
    "M": 10,  # Votes
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Dark divider column F
for row in range(1, CB_END + 10):
    ws.cell(row, 6).fill = PatternFill("solid", fgColor=BLUE_DARK)

ws.freeze_panes = "G5"
ws.sheet_properties.tabColor = GOLD

# ============================================================
# Sheet: By District
# ============================================================
ws_bd = wb.create_sheet("By District")
ws_bd.merge_cells("A1:H1")
ws_bd["A1"].value     = "Indiana 2026 Primary Winners — All Districts"
ws_bd["A1"].font      = Font(bold=True, size=14, color=WHITE)
ws_bd["A1"].fill      = PatternFill("solid", fgColor=BLUE_DARK)
ws_bd["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_bd.row_dimensions[1].height = 30

bd_hdrs = ["County","District","Type","Dist #","Candidate Name","Party","Primary Result","Votes"]
ws_bd.append(bd_hdrs)
for col, hdr in enumerate(bd_hdrs, 1):
    c = ws_bd.cell(2, col)
    c.font = Font(bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE_MID)
    c.alignment = Alignment(horizontal="center")
    c.border = thin()
ws_bd.row_dimensions[2].height = 20

sorted_rows = sorted(rows, key=lambda r: (r[1], r[3], r[8], -r[9], r[0], r[5]))
for idx, r in enumerate(sorted_rows, 3):
    county, _ord, dtype, dnum, dlabel, cand, party, result, rsort, votes = r
    bg = GREY_LIGHT if idx % 2 == 0 else WHITE
    for col, val in enumerate([county, dlabel, dtype, dnum, cand, party, result, votes], 1):
        cell = ws_bd.cell(idx, col)
        cell.value = val
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(vertical="center")
        cell.border = thin()
    pc = ws_bd.cell(idx, 6)
    if party == "D":
        pc.fill = PatternFill("solid", fgColor=RED_LIGHT); pc.font = Font(color="8B0000", bold=True, size=10)
    elif party == "R":
        pc.fill = PatternFill("solid", fgColor=BLUE_LIGHT); pc.font = Font(color="00008B", bold=True, size=10)
    else:
        pc.fill = PatternFill("solid", fgColor=GREEN_LIGHT); pc.font = Font(color="005500", bold=True, size=10)
    rc = ws_bd.cell(idx, 7)
    if "Won" in result:
        rc.fill = PatternFill("solid", fgColor=WIN_FILL); rc.font = Font(color="155724", bold=True, size=10)

for col, w in zip("ABCDEFGH", [18,10,15,8,40,8,18,12]):
    ws_bd.column_dimensions[col].width = w
ws_bd.freeze_panes = "A3"
ws_bd.auto_filter.ref = f"A2:H{len(sorted_rows)+2}"
ws_bd.sheet_properties.tabColor = BLUE_MID

# ============================================================
# Sheet: Apps Script
# ============================================================
ws_as = wb.create_sheet("Apps Script")
ws_as.merge_cells("A1:G1")
ws_as["A1"].value     = "Google Apps Script — Setup Instructions + Full Code"
ws_as["A1"].font      = Font(bold=True, size=13, color=WHITE)
ws_as["A1"].fill      = PatternFill("solid", fgColor=BLUE_DARK)
ws_as["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_as.row_dimensions[1].height = 28

notes = [
    ("A3",  "WHAT THIS SCRIPT DOES:"),
    ("A4",  "• onSelectionChange — clicking a county name toggles it selected/unselected"),
    ("A5",  "• onEdit — toast notification when filters change"),
    ("A6",  "• IRS Tools menu — Select All / Clear All / Reset District Filter"),
    ("A8",  "HOW TO INSTALL (one-time setup):"),
    ("A9",  "1. Extensions → Apps Script"),
    ("A10", "2. Delete any placeholder code, paste ALL code from row 17 onward"),
    ("A11", "3. Save (Ctrl+S / Cmd+S), then close the Apps Script tab"),
    ("A12", "4. Reload the spreadsheet — IRS Tools menu appears in the menu bar"),
    ("A13", "5. onSelectionChange activates immediately; clicking county names now toggles them"),
    ("A15", "── APPS SCRIPT CODE BELOW ──"),
]
for addr, text in notes:
    cell = ws_as[addr]
    cell.value = text
    cell.alignment = Alignment(wrap_text=True)
    cell.font = (Font(bold=True, size=11, color=BLUE_DARK) if addr in ("A3","A8")
                 else Font(bold=True, size=10, color=GOLD) if addr == "A15"
                 else Font(size=10))

apps_script = r"""
// ============================================================
// Paste this entire block into Extensions → Apps Script
// ============================================================

// ---- Custom menu ----
function onOpen() {
  SpreadsheetApp.getUi()
    .createMenu('IRS Tools')
    .addItem('Select All Counties',   'selectAllCounties')
    .addItem('Clear All Counties',    'clearAllCounties')
    .addSeparator()
    .addItem('Reset District Filter', 'resetDistrictFilter')
    .addSeparator()
    .addItem('Go to By District tab', 'goToByDistrict')
    .addToUi();
}

// ---- Click a county name to toggle selection ----
// County name cells: col B (2) rows 6-51 → state in col C (3)
//                   col D (4) rows 6-51 → state in col E (5)
function onSelectionChange(e) {
  const sheet = e.range.getSheet();
  if (sheet.getName() !== 'Lookup') return;

  // Only act on single-cell clicks (ignore drag selections)
  if (e.range.getNumRows() !== 1 || e.range.getNumColumns() !== 1) return;

  const col = e.range.getColumn();
  const row = e.range.getRow();
  if (row < 6 || row > 51) return;

  let stateCol = null;
  if (col === 2) stateCol = 3;   // clicked left county name → toggle col C
  if (col === 4) stateCol = 5;   // clicked right county name → toggle col E
  if (!stateCol) return;

  const stateCell = sheet.getRange(row, stateCol);
  const cur = stateCell.getValue();
  // Handle both boolean true AND text "TRUE" (xlsx import stores text)
  const isSelected = (cur === true || cur === 'TRUE');
  stateCell.setValue(!isSelected);  // always write boolean
}

// ---- Toast feedback when filters change ----
function onEdit(e) {
  const sheet = e.range.getSheet();
  if (sheet.getName() !== 'Lookup') return;

  const col  = e.range.getColumn();
  const row  = e.range.getRow();
  const ss   = e.source;

  // State cells (C or E, rows 6-51) updated by onSelectionChange
  if ((col === 3 || col === 5) && row >= 6 && row <= 51) {
    const isSel = v => v === true || v === 'TRUE';
    const n = sheet.getRange('C6:C51').getValues().flat().filter(isSel).length
            + sheet.getRange('E6:E51').getValues().flat().filter(isSel).length;
    ss.toast(
      n === 0 ? 'All 92 counties shown.' : n + ' ' + (n === 1 ? 'county' : 'counties') + ' selected.',
      'County Filter', 3
    );
    return;
  }

  // District type dropdown (I4) or district number input (L4)
  const addr = e.range.getA1Notation();
  if (addr === 'I4' || addr === 'L4') {
    const type = sheet.getRange('I4').getValue() || 'All Types';
    const num  = sheet.getRange('L4').getValue();
    ss.toast(
      'Showing: ' + type + (num ? ', District ' + num : ', all districts'),
      'District Filter', 3
    );
  }
}

// ---- Bulk county helpers ----
function selectAllCounties() {
  const lu = SpreadsheetApp.getActiveSpreadsheet().getSheetByName('Lookup');
  lu.getRange('C6:C51').setValue(true);
  lu.getRange('E6:E51').setValue(true);
  SpreadsheetApp.getActiveSpreadsheet().toast('All 92 counties selected.', 'IRS Tools', 3);
}

function clearAllCounties() {
  const lu = SpreadsheetApp.getActiveSpreadsheet().getSheetByName('Lookup');
  lu.getRange('C6:C51').setValue(false);
  lu.getRange('E6:E51').setValue(false);
  SpreadsheetApp.getActiveSpreadsheet().toast('Showing all 92 counties.', 'IRS Tools', 3);
}

// ---- Reset district filter ----
function resetDistrictFilter() {
  const lu = SpreadsheetApp.getActiveSpreadsheet().getSheetByName('Lookup');
  lu.getRange('I4').setValue('All Types');
  lu.getRange('L4').clearContent();
  SpreadsheetApp.getActiveSpreadsheet().toast('District filter cleared.', 'IRS Tools', 2);
}

function goToByDistrict() {
  const ss = SpreadsheetApp.getActiveSpreadsheet();
  ss.setActiveSheet(ss.getSheetByName('By District'));
}
"""

for line_idx, line in enumerate(apps_script.strip().split("\n"), 17):
    cell = ws_as.cell(line_idx, 1)
    cell.value = line
    cell.font  = Font(name="Courier New", size=9, color="003300")

ws_as.column_dimensions["A"].width = 90
ws_as.sheet_properties.tabColor    = "888888"

# ============================================================
# Save
# ============================================================
out_path = os.path.join(base, "Indiana_County_Candidates_2026.xlsx")
wb.save(out_path)
print(f"Saved  : {out_path}")
print(f"Rows   : {len(rows)}")
print(f"Tabs   : {[s.title for s in wb.worksheets]}")
